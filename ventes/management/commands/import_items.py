"""
Commande Django management pour importer les soumissions

Placer ce fichier dans: votre_app/management/commands/import_soumissions.py

Usage:
    python manage.py import_soumissions soumissions_nom.xlsx
    python manage.py import_soumissions soumissions_nom.xlsx --update  # Met à jour les existantes
"""

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from django.contrib.auth import get_user_model
from ventes.models import Soumission
from datetime import datetime

User = get_user_model()


class Command(BaseCommand):
    help = 'Importe les soumissions depuis un fichier Excel et les lie aux vendeurs'

    def add_arguments(self, parser):
        parser.add_argument(
            'fichier',
            type=str,
            help='Chemin vers le fichier Excel à importer'
        )
        parser.add_argument(
            '--update',
            action='store_true',
            help='Met à jour les soumissions existantes si elles existent déjà'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Simule l\'importation sans modifier la base de données'
        )
        parser.add_argument(
            '--stats',
            action='store_true',
            help='Affiche uniquement les statistiques du fichier sans importer'
        )
        parser.add_argument(
            '--debug',
            action='store_true',
            help='Affiche les 5 premières lignes avec détails pour déboguer'
        )

    def parse_date(self, date_value):
        """Parse une date depuis différents formats"""
        if not date_value:
            return None
        
        # Si c'est déjà un objet datetime
        if isinstance(date_value, datetime):
            return date_value.date()
        
        # Si c'est déjà un objet date
        from datetime import date
        if isinstance(date_value, date):
            return date_value
        
        # Essayer plusieurs formats de date
        date_formats = [
            '%Y-%m-%d',      # 2025-12-12
            '%d/%m/%Y',      # 12/12/2025
            '%d-%m-%Y',      # 12-12-2025
            '%m/%d/%Y',      # 12/12/2025 (US)
            '%Y/%m/%d',      # 2025/12/12
        ]
        
        date_str = str(date_value).strip()
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except (ValueError, TypeError):
                continue
        
        # Si rien ne fonctionne, afficher pour debug
        self.stdout.write(self.style.WARNING(f'⚠️  Format de date non reconnu: "{date_value}"'))
        return None

    def parse_datetime(self, datetime_value):
        """Parse un datetime depuis différents formats"""
        if not datetime_value:
            return None
        
        if isinstance(datetime_value, datetime):
            return datetime_value
        
        try:
            return datetime.strptime(str(datetime_value), '%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError):
            return None

    def bool_from_string(self, value):
        """Convertit une valeur en booléen"""
        if value in [True, 1, '1', 'True', 'true', 'TRUE']:
            return True
        return False

    def normaliser_statut(self, status):
        """Normalise le statut en enlevant les accents et en le mettant en minuscules"""
        if not status:
            return 'en_cours'
        
        # Convertir en string et nettoyer
        status_str = str(status).strip().lower()
        
        # Mapping des statuts avec et sans accents
        statut_mapping = {
            'en_cours': 'en_cours',
            'en cours': 'en_cours',
            'envoye': 'envoye',
            'envoyé': 'envoye',
            'envoyé': 'envoye',  # Avec accent aigu
            'accepte': 'accepte',
            'accepté': 'accepte',
            'refuse': 'refuse',
            'refusé': 'refuse',
        }
        
        return statut_mapping.get(status_str, 'en_cours')

    def afficher_statistiques(self, sheet):
        """Affiche les statistiques du fichier Excel"""
        self.stdout.write('\n' + '='*60)
        self.stdout.write(self.style.SUCCESS('STATISTIQUES DU FICHIER'))
        self.stdout.write('='*60)
        
        users_count = {}
        with_date = 0
        without_date = 0
        total_rows = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            user_identifier = row[0]
            date_evenement = self.parse_date(row[12] if len(row) > 12 else None)  # ✅ CORRIGÉ: index 12
            
            if not user_identifier and not date_evenement:
                continue  # Ligne vide
            
            total_rows += 1
            
            if user_identifier:
                user_str = str(user_identifier)
                users_count[user_str] = users_count.get(user_str, 0) + 1
            
            if date_evenement:
                with_date += 1
            else:
                without_date += 1
        
        self.stdout.write(f'📊 Total de lignes: {total_rows}')
        self.stdout.write(f'✅ Lignes avec date: {with_date}')
        self.stdout.write(f'⚠️  Lignes sans date: {without_date}')
        
        self.stdout.write('\n👥 Utilisateurs trouvés dans le fichier:')
        for user, count in sorted(users_count.items()):
            # Vérifier si l'utilisateur existe
            exists = self.get_user_by_identifier(user) is not None
            status = '✓' if exists else '✗'
            self.stdout.write(f'   {status} "{user}": {count} soumissions')
        
        self.stdout.write('='*60)
        self.stdout.write('💡 Utilisez --dry-run pour simuler l\'import')
        self.stdout.write('='*60 + '\n')

    def afficher_debug(self, sheet):
        """Affiche les détails des premières lignes pour déboguer"""
        self.stdout.write('\n' + '='*60)
        self.stdout.write(self.style.SUCCESS('MODE DEBUG - DÉTAILS DES 5 PREMIÈRES LIGNES'))
        self.stdout.write('='*60)
        
        # Afficher les en-têtes
        headers = [cell.value for cell in sheet[1]]
        self.stdout.write('\n📋 En-têtes du fichier:')
        for i, header in enumerate(headers):
            self.stdout.write(f'   Colonne {i}: {header}')
        
        self.stdout.write('\n📊 Contenu COMPLET des 3 premières lignes:')
        for index, row in enumerate(sheet.iter_rows(min_row=2, max_row=4, values_only=True), start=2):
            self.stdout.write(f'\n{"="*50}')
            self.stdout.write(f'LIGNE {index}')
            self.stdout.write(f'{"="*50}')
            for i, value in enumerate(row):
                if value is not None and value != '':
                    header = headers[i] if i < len(headers) else f'Col{i}'
                    self.stdout.write(f'  [{i:2d}] {header:20s} = {value} (type: {type(value).__name__})')
        
        self.stdout.write('\n' + '='*60 + '\n')

    def get_user_by_identifier(self, user_identifier):
        """Trouve un utilisateur par nom d'utilisateur uniquement"""
        if not user_identifier:
            return None
        
        try:
            # Convertir en string et chercher uniquement par username
            user_str = str(user_identifier).strip()
            
            # Ignorer les IDs numériques purs (comme 1, 2, 3)
            if user_str.isdigit():
                self.stdout.write(
                    self.style.WARNING(f'⚠️  Valeur numérique "{user_str}" ignorée - username attendu')
                )
                return None
            
            # Chercher par username (insensible à la casse)
            return User.objects.filter(username__iexact=user_str).first()
        except Exception as e:
            self.stdout.write(self.style.WARNING(f'⚠️  Erreur recherche utilisateur: {e}'))
            return None

    def get_or_create_numero_soumission(self, index):
        """Génère un numéro de soumission unique"""
        return f"SOU-{index:05d}"

    def handle(self, *args, **options):
        fichier = options['fichier']
        update = options['update']
        dry_run = options['dry_run']
        show_stats = options['stats']
        debug = options['debug']
        
        try:
            workbook = openpyxl.load_workbook(fichier)
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f'❌ Fichier non trouvé: {fichier}'))
            return
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Erreur lors de la lecture du fichier: {str(e)}'))
            return
        
        sheet = workbook.active
        
        # Si option --debug, afficher les détails des premières lignes
        if debug:
            self.afficher_debug(sheet)
            return
        
        # Si option --stats, afficher uniquement les statistiques
        if show_stats:
            self.afficher_statistiques(sheet)
            return
        
        if dry_run:
            self.stdout.write(self.style.WARNING('🔍 MODE DRY-RUN - Aucune modification ne sera effectuée'))
        
        stats = {
            'soumissions_creees': 0,
            'soumissions_mises_a_jour': 0,
            'erreurs': 0,
            'utilisateurs_non_trouves': set(),
        }
        
        def traiter_importation():
            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                try:
                    user_identifier = row[0]  # user (colonne A)
                    nom_compagnie = row[1] if len(row) > 1 else None  # company_name (colonne B)
                    adresse = row[2] if len(row) > 2 else None  # event_location (colonne C)
                    commande_par = row[3] if len(row) > 3 else None  # ordered_by (colonne D)
                    telephone = row[4] if len(row) > 4 else None  # phone (colonne E)
                    email = row[5] if len(row) > 5 else None  # email (colonne F)
                    # row[6] = billing_address (ignoré pour l'instant)
                    avec_service = self.bool_from_string(row[7] if len(row) > 7 else False)  # avec_service (H)
                    # row[8] = avec_service_md (ignoré)
                    avec_alcool = self.bool_from_string(row[9] if len(row) > 9 else False)  # avec_alcool (J)
                    # row[10] = language (ignoré)
                    location_materiel = self.bool_from_string(row[11] if len(row) > 11 else False)  # location_materiel (L)
                    date_evenement = self.parse_date(row[12] if len(row) > 12 else None)  # date (M) ✅ CORRIGÉ
                    nombre_personnes = row[13] if len(row) > 13 else 0  # guest_count (N)
                    created_at = self.parse_datetime(row[14] if len(row) > 14 else None)  # created_at (O)
                    updated_at = self.parse_datetime(row[15] if len(row) > 15 else None)  # updated_at (P)
                    sent_at = self.parse_datetime(row[16] if len(row) > 16 else None)  # sent_at (Q)
                    status = row[18] if len(row) > 18 else 'en_cours'  # status (S)
                    
                    if not user_identifier or str(user_identifier).strip() == '':
                        self.stdout.write(
                            self.style.WARNING(f'⚠️  Ligne {index + 1}: utilisateur manquant, ignorée')
                        )
                        continue
                    
                    utilisateur = self.get_user_by_identifier(user_identifier)
                    if not utilisateur:
                        stats['utilisateurs_non_trouves'].add(str(user_identifier))
                        self.stdout.write(
                            self.style.WARNING(
                                f'⚠️  Ligne {index + 1}: Utilisateur "{user_identifier}" non trouvé - soumission ignorée'
                            )
                        )
                        continue  # ✅ Ignore complètement cette ligne
                    
                    # Si pas de date ET pas de données utiles, ignorer
                    if not date_evenement and not nom_compagnie:
                        continue  # Ligne vide, ignorer silencieusement
                    
                    if not date_evenement:
                        self.stdout.write(
                            self.style.WARNING(f'⚠️  Ligne {index + 1}: date manquante, ignorée')
                        )
                        continue
                    
                    numero_soumission = self.get_or_create_numero_soumission(index)
                    
                    soumission_data = {
                        'nom_compagnie': nom_compagnie or '',
                        'adresse': adresse or '',
                        'commande_par': commande_par or '',
                        'telephone': telephone or '',
                        'email': email or '',
                        'avec_service': avec_service,
                        'avec_alcool': avec_alcool,
                        'location_materiel': location_materiel,
                        'date_evenement': date_evenement,
                        'nombre_personnes': int(nombre_personnes) if nombre_personnes else 0,
                        'statut': self.normaliser_statut(status),
                        'cree_par': utilisateur,
                    }
                    
                    if created_at:
                        soumission_data['cree_a'] = created_at
                    if updated_at:
                        soumission_data['date_modification'] = updated_at
                    if sent_at and soumission_data['statut'] == 'envoye':
                        soumission_data['date_soumission'] = sent_at
                    
                    if not dry_run:
                        if update:
                            soumission, created = Soumission.objects.update_or_create(
                                numero_soumission=numero_soumission,
                                defaults=soumission_data
                            )
                        else:
                            try:
                                soumission = Soumission.objects.create(
                                    numero_soumission=numero_soumission,
                                    **soumission_data
                                )
                                created = True
                            except Exception:
                                created = False
                                self.stdout.write(
                                    self.style.WARNING(
                                        f'⚠️  Ligne {index + 1}: Soumission {numero_soumission} existe déjà'
                                    )
                                )
                                continue
                    else:
                        created = not Soumission.objects.filter(
                            numero_soumission=numero_soumission
                        ).exists()
                    
                    if created:
                        stats['soumissions_creees'] += 1
                    else:
                        stats['soumissions_mises_a_jour'] += 1
                    
                    if (index) % 50 == 0:
                        self.stdout.write(f'📊 Progression: {index} lignes traitées...')
                        
                except Exception as e:
                    stats['erreurs'] += 1
                    self.stdout.write(
                        self.style.ERROR(f'❌ Erreur ligne {index + 1}: {str(e)}')
                    )
                    continue
        
        if not dry_run:
            with transaction.atomic():
                traiter_importation()
        else:
            traiter_importation()
        
        self.stdout.write('\n' + '='*60)
        self.stdout.write(self.style.SUCCESS('RAPPORT D\'IMPORTATION DES SOUMISSIONS'))
        self.stdout.write('='*60)
        self.stdout.write(f'✅ Soumissions créées: {stats["soumissions_creees"]}')
        self.stdout.write(f'🔄 Soumissions mises à jour: {stats["soumissions_mises_a_jour"]}')
        self.stdout.write(self.style.ERROR(f'❌ Erreurs: {stats["erreurs"]}'))
        self.stdout.write(f'📊 Total traité: {stats["soumissions_creees"] + stats["soumissions_mises_a_jour"]}')
        
        if stats['utilisateurs_non_trouves']:
            self.stdout.write('\n⚠️  Utilisateurs non trouvés (vous devez créer ces comptes):')
            for user_id in sorted(stats['utilisateurs_non_trouves']):
                self.stdout.write(f'   - "{user_id}"')
            self.stdout.write('\n💡 Pour créer les utilisateurs manquants:')
            self.stdout.write('   python manage.py createsuperuser --username <nom>')
        
        self.stdout.write('='*60)
        
        if dry_run:
            self.stdout.write(self.style.WARNING('\n🔍 Dry-run terminé - Aucune modification effectuée'))